
# Import từ chính thư mục utils
from . import utils_functions
from . import utils_models
from . import define

import traceback
from datetime import datetime

import os
import sys
import time
import inspect
import tkinter as tk
from tkinter import ttk
from tkinter import font
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import xlwings as xw

import pyodbc
import json
from cryptography.fernet import Fernet
import datetime
from datetime import datetime
from decimal import Decimal
from PIL import Image, ImageTk
import pandas as pd
import re
import math

class utils_controller_action_after_event_250216_14h57:
    
    def f_get_the_latest_number_of_slip(entry_so_phieu, ma_thanh_vien, loai_phieu):
        # Get the latest number of slip
        so_phieu = utils_controller_get_the_latest_number_of_slip.handle_button_get_number_of_slip_click()
        
        # Create the connection string
        connection_number_of_slip = f"{ma_thanh_vien}-{loai_phieu}-{so_phieu + 1}"
        
        # Config the entry_so_phieu
        entry_so_phieu.config(state="normal")
        entry_so_phieu.delete(0, tk.END)
        entry_so_phieu.insert(0, connection_number_of_slip)
        entry_so_phieu.config(state="readonly")

    def update_entry_id_after_adding_new_row(tree, entry_id):
        row_count = 1 + len(tree.get_children())    
        entry_id.config(state="normal")  # Enable the Entry widget to update the value
        entry_id.delete(0, tk.END)  # Clear the existing value
        entry_id.insert(0, row_count)  # Insert the new value (ID)
        entry_id.config(state="disabled")  # Disable the Entry widget again

class utils_controller_config_notification_250220_10h05:
    def f_config_notification(element_label, text="...", fg="blue"):
            element_label.config(text=text, fg=fg)

class utils_controller_Export_data_to_Excel_250222_09h16:
    
    def export_log_to_excel(query, my_excel_header):
        # Truyền câu query vào model để lấy dữ liệu từ SQL (Giả sử hàm này trả về DataFrame)
        df = utils_models.utils_model_get_data_from_SQL.get_data_with_query(query)
        
        # Kiểm tra nếu có dữ liệu trả về
        if isinstance(df, list) and df:
            path = utils_controller_Export_data_to_Excel_250222_09h16.f_utils_export_data_to_excel(my_excel_header, df)
            if path:
                return True, path
        return False, None
    
    def f_utils_export_data_to_excel(data_header, data):
        # Kiểm tra tính nhất quán của dữ liệu
        if not data or not data_header:
            print("No data or headers provided.")
            return
        
        # Chuyển đổi dữ liệu từ pyodbc.Row thành tuple (nếu cần)
        data = [tuple(row) for row in data]

        # Kiểm tra số lượng cột có khớp không
        if len(data[0]) != len(data_header):
            print(f"Error: Mismatch in column count. Expected {len(data_header)}, but got {len(data[0])}.")
            return
        
        # Tạo DataFrame
        df = pd.DataFrame(data, columns=data_header)
        
        # Chuyển đổi kiểu dữ liệu Decimal và datetime để tương thích với Excel
        for column in df.columns:
            df[column] = df[column].apply(lambda x: float(x) if isinstance(x, Decimal) 
                                        else x.strftime('%Y-%m-%d %H:%M:%S') if isinstance(x, datetime) 
                                        else x)
        
        # Đường dẫn thư mục lưu file
        directory = define.PATH_DEFAUL
        if not os.path.exists(directory):
            os.makedirs(directory)  # Tạo thư mục nếu chưa tồn tại
        
        # Tạo file với tên không trùng
        file_path = utils_functions.f_utils_get_unique_filename(directory, "exported_data.xlsx")

        # Xuất ra file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Exported Data"
        
        # Ghi header
        ws.append(data_header)
        
        # Ghi dữ liệu
        for row in df.itertuples(index=False, name=None):
            ws.append(row)
        
        # Lưu file
        wb.save(file_path)
        wb.close()
        # print(f"Data exported successfully to {file_path}")
        utils_controller_Export_data_to_Excel_250222_09h16.format_excel_file(file_path, ws.title, "A1")
        return file_path
    
    def format_excel_file(file_path, sheet_name, start_cell):
        # Mở file Excel
        wb = openpyxl.load_workbook(file_path)
        
        # Kiểm tra nếu sheet tồn tại
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' không tồn tại trong {file_path}")
            return
        
        ws = wb[sheet_name]  # Chọn sheet theo tên
        
        # Xác định vị trí bắt đầu (ví dụ: "A1")
        start_row = ws[start_cell].row
        start_col = ws[start_cell].column

        # Tìm hàng và cột lớn nhất có dữ liệu
        max_row = ws.max_row
        max_col = ws.max_column

        # Xác định phạm vi dữ liệu (Dynamic)
        range_to_format = ws[start_cell:get_column_letter(max_col) + str(max_row)]

        # Áp dụng định dạng cho từng ô
        for row in range_to_format:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",  # Căn giữa ngang
                    vertical="center",    # Căn giữa dọc
                    wrap_text=False,      # Không xuống dòng tự động
                    shrink_to_fit=False   # Không co chữ vừa ô
                )

        # Tự động điều chỉnh độ rộng cột (AutoFit trong VBA)
        for col_num in range(start_col, max_col + 1):
            col_letter = get_column_letter(col_num)  # Lấy chữ cái của cột (A, B, C,...)
            max_length = 0

            for row_num in range(start_row, max_row + 1):
                cell_value = ws[f"{col_letter}{row_num}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws.column_dimensions[col_letter].width = max_length + 2  # Thêm padding

        # Lưu file
        wb.save(file_path)
        wb.close()


class utils_controller_get_the_latest_number_of_slip:
    
    def get_list_number_of_slip(database_name, table_name, slip_column_name):
        # Tạo câu query SQL với danh sách số phiếu
        query = f"""
        SELECT DISTINCT
            {slip_column_name}
        FROM [{database_name}].[dbo].[{table_name}]
        WHERE [XOA_SUA] = ''
        """
        # print("query", query)
        
        # lấy danh sách số phiếu từ SQL
        danh_sach_so_phieu = utils_models.utils_model_get_data_from_SQL.get_data_with_query(query)
        # print("danh_sach_so_phieu", danh_sach_so_phieu)
        
        return danh_sach_so_phieu
        
    def extract_numbers_from_data_SQL_num_01(data):
        data_01 = data
        data_02 = tuple(int(item[0].split('-')[-1]) for item in data_01)
        if not data_02:
            current_year = str(datetime.now().year)[-2:]  # Lấy 2 số cuối của năm hiện tại
            data_final = int(f'{current_year}0000')
        else:
            data_final = max(data_02)
        return data_final
        
    def handle_button_get_number_of_slip_click(database_name, table_name):
        # Lấy danh sách số phiếu từ SQL
        data_01 = utils_controller_get_the_latest_number_of_slip.get_list_number_of_slip(database_name, table_name)
        # Lấy số phiếu cuối cùng
        data_02 = utils_controller_get_the_latest_number_of_slip.extract_numbers_from_data_SQL_num_01(data_01)
        
        return data_02
   
class utils_controller_get_the_header_of_table_in_SQL_250221_11h01:
    def get_column_names(database_name, table_name):
        try:
            query = f"""
            SELECT COLUMN_NAME
            FROM [{database_name}].INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = '{table_name}' AND TABLE_SCHEMA = 'dbo'
            """
            # lấy danh sách tên cột từ SQL
            column_headers = utils_models.utils_model_get_data_from_SQL.get_data_with_query(query)
            if not column_headers:
                return None
            
            return column_headers
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return None
             
class utils_controller_set_size_of_windown_250215_10h24:    
    def f_utils_set_window_size_of_new_view(root, width=0, height=0, maximize=False):
        """Set the window size to 4/5 of the screen size or maximize it."""
        if maximize:
            root.state('zoomed')  # Maximizes the window (Windows)
            # root.is_maximized = True  # Set the flag to track maximized state
            # root.bind("<Configure>", utils_controller_set_size_of_windown_250215_10h24.on_resize)  # Bind the resize event
            return

        if width == 0 or height == 0:
            # Get screen dimensions
            screen_width = root.winfo_screenwidth()
            screen_height = root.winfo_screenheight()
            
            # Calculate 4/5 of screen dimensions
            ratio = 0.75
            height = int(screen_height * ratio)
            width = int(screen_width * ratio)
        
        # Set the window size
        root.geometry(f"{width}x{height}")

    # def on_resize(event):
    #     """Called when the window is resized, adjusts the window to 75% of the screen if maximized."""
    #     if isinstance(event.widget, tk.Tk):  # Ensure we're working with the root window
    #         root = event.widget
            
    #         # # Only apply the resizing logic if the window was maximized
    #         # if getattr(root, 'is_maximized', False):  # Check if window was maximized
    #         #     screen_width = root.winfo_screenwidth()
    #         #     screen_height = root.winfo_screenheight()

    #         #     # Calculate 75% of screen dimensions
    #         #     ratio = 0.75
    #         #     new_width = int(screen_width * ratio)
    #         #     new_height = int(screen_height * ratio)

    #         #     current_width = root.winfo_width()
    #         #     current_height = root.winfo_height()

    #         #     # Only update the window size if it's different from the current size
    #         #     if current_width != new_width or current_height != new_height:
    #         #         root.geometry(f"{new_width}x{new_height}")
    #         #         root.is_maximized = False
    #         # else:
    #         #     # If the window is not maximized, do nothing
    #         #     pass
            
    #         if root.winfo_width() < 1000 or root.winfo_height() < 750:
    #             root.geometry("1000x750")
            
class utils_controller_TreeviewConfigurator_250217_13h20:
    """Class để áp dụng cấu hình cho Treeview"""

    @staticmethod
    def apply_treeview_config(my_treeview, config_json_path):
        """Cấu hình Treeview dựa trên JSON"""
        
        # Xóa các cột hiện tại
        my_treeview.delete(*my_treeview.get_children())
        for col in my_treeview["columns"]:
            my_treeview.heading(col, text="")  # Xóa tiêu đề

        # Load dữ liệu từ JSON
        data = utils_models.utils_model_TreeviewConfigLoader_250217_13h20.load_json(config_json_path)
        if not data:
            return
        
        columns_config, column_names = utils_models.utils_model_TreeviewConfigLoader_250217_13h20.get_column_config(data)
        my_treeview["columns"] = column_names

        # Lấy cấu hình font header
        header_font = utils_models.utils_model_TreeviewConfigLoader_250217_13h20.get_header_font(data)

        # Cấu hình cột
        for config, col in zip(columns_config, column_names):
            my_treeview.heading(col, text=col)
            my_treeview.column(
                col,
                width=config.get("width", 100),
                minwidth=config.get("min_width", 50),
                anchor=config.get("anchor", "w"),
                stretch=config.get("stretch", True)
            )

        # Cấu hình font tiêu đề
        style = ttk.Style()
        style.configure("Treeview.Heading", font=header_font)

        # print("Treeview đã được cấu hình thành công từ JSON.")

class utils_controller_TreeviewHandler_click_250217_22h34:

    def treeview_single_click(my_Treeview):
        selected_item = my_Treeview.selection()
        if selected_item:
            row_values = my_Treeview.item(selected_item[0], "values")
            result_tuple = tuple(row_values) if row_values else None
            # print(result_tuple)
            return result_tuple
        return None

    def treeview_double_click(my_Treeview, column_return):
        selected_item = my_Treeview.selection()
        if selected_item:
            row_values = my_Treeview.item(selected_item[0], "values")
            return row_values[column_return] if len(row_values) > column_return else None
        return None
    
class utils_controller_get_information_of_database:
        
    def load_id_nhan_vien():
        ID_nhan_vien = "NV01"
        return ID_nhan_vien
    
    def load_xoa_sua_mac_dinh():
        Xoa_Sua = ""
        return Xoa_Sua
        
    def load_expired_mac_dinh():
        Expired = 0
        return Expired
        
    def load_danh_sach_id_duoc_phan_quyen():
        danh_sach_id_duoc_phan_quyen = "'NV01', 'NV02', 'NV03'"
        return danh_sach_id_duoc_phan_quyen
        
    def load_ma_thanh_vien():
        ma_thanh_vien = "TB"
        return ma_thanh_vien
    
    def load_database_name():
        database_name = "TBD_2024"
        return database_name
    
    def load_table_name_TB_KD02_KE_HOACH_DAT_HANG():
        table_name = "TB_KD02_KE_HOACH_DAT_HANG"
        return table_name
    
    def load_table_name_TB_KD02_YEU_CAU_DAT_HANG():
        table_name = "TB_KD02_YEU_CAU_DAT_HANG"
        return table_name
    
    def load_table_name_TB_KD03_DE_NGHI_XUAT_KHO():
        table_name = "TB_KD03_DE_NGHI_XUAT_KHO"
        return table_name
    
    def load_table_name_TB_AD00_DANH_SACH_KHACH_HANG():
        table_name = "TB_AD00_DANH_SACH_KHACH_HANG"
        return table_name

    def load_table_name_TB_INVENTORY_CATEGORIES():
        table_name = "TB_INVENTORY_CATEGORIES_250214_09h05"
        return table_name
    
    def load_table_name_VIEW_QUYET_TOAN_YEU_CAU_DAT_HANG():
        table_name = "VIEW_QUYET_TOAN_YEU_CAU_DAT_HANG"
        return table_name
    
class utils_controller_get_data_from_SQL_to_treeview_with_quey_and_params_list:
    def load_data_to_treeview(my_treeview, data):
        # Lấy số lượng cột từ my_treeview
        columns_count = len(my_treeview["columns"])
        
        # Xóa hết các item trong treeview hiện tại
        for item in my_treeview.get_children():
            my_treeview.delete(item)
            
        # Duyệt qua từng dòng dữ liệu và chèn vào treeview
        for row in data:
            # Chỉ lấy các giá trị trong row tương ứng với số cột
            values = row[:columns_count]  # Cắt bỏ các cột thừa
            
            # Kiểm tra và xử lý kiểu dữ liệu nếu cần thiết (chẳng hạn ngày tháng)
            processed_values = []
            for value in values:
                if isinstance(value, str):  # Nếu là chuỗi, giữ nguyên
                    processed_values.append(value)
                elif isinstance(value, (int, float)):  # Nếu là số, giữ nguyên
                    processed_values.append(value)
                elif isinstance(value, datetime):  # Nếu là ngày tháng, chuyển sang dạng chuỗi
                    processed_values.append(value.strftime("%d-%m-%Y"))
                else:
                    processed_values.append(value)  # Thêm giá trị gốc nếu không thuộc các loại trên

            # Nếu thiếu cột, bổ sung None
            if len(processed_values) < columns_count:
                processed_values.extend([None] * (columns_count - len(processed_values)))
            
            # Chèn dữ liệu vào Treeview
            my_treeview.insert("", "end", values=processed_values)
        
    def load_data_with_quey_and_params(my_treeview, query, params_list):
        data = utils_models.utils_model_SQL_server.fetch_data_with_quey_and_params(query, params_list)
        utils_controller_get_data_from_SQL_to_treeview_with_quey_and_params_list.load_data_to_treeview(my_treeview, data)    
    
class utils_controller_check_exist:
    def check_exist_values(values_to_check, database_name, table_name, column_name):
        # Nếu đầu vào là None hoặc rỗng, trả về False
        if not values_to_check:
            notification_text = f"Không có giá trị nào để kiểm tra."
            return False, notification_text

        # Nếu chỉ có một giá trị (chuỗi hoặc số), chuyển thành tuple
        if isinstance(values_to_check, (str, int, float)):
            values_to_check = (values_to_check,)
        # Nếu là danh sách, chuyển thành tuple
        elif isinstance(values_to_check, list):
            values_to_check = tuple(values_to_check)
        # Nếu không phải tuple sau khi kiểm tra, return False
        elif not isinstance(values_to_check, tuple):
            notification_text = f"Các giá trị cần kiểm tra không ở định dạng phù hợp: {values_to_check}"
            return False, notification_text

        try:
            conn = utils_functions.f_utils_create_a_connection_string_to_SQL_Server()
            cursor = conn.cursor()
            
            placeholders = ', '.join('?' * len(values_to_check))
            query = (f"SELECT {column_name} FROM {database_name}.[dbo].{table_name} "
                    f"WHERE {column_name} IN ({placeholders}) AND [XOA_SUA] = ''")
            
            cursor.execute(query, values_to_check)
            existing_values = {row[0] for row in cursor.fetchall()}
            
            cursor.close()
            conn.close()
            
            missing_values = set(values_to_check) - existing_values
            
            if not missing_values:
                notification_text = ""
                return True, notification_text
            else:
                notification_text = f"Các giá trị chưa tồn tại: {missing_values}"
                return False, notification_text
        
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            notification_text = ""
            return False, notification_text
        
    def get_unique_column_values(my_treeview, column_index):
        """
        Lấy danh sách giá trị không trùng lặp từ một cột trong Treeview.

        :param treeview: Đối tượng Treeview
        :param column_index: Vị trí cột (bắt đầu từ 0)
        :return: Danh sách các giá trị không trùng lặp
        """
        unique_values = set()
        for child in my_treeview.get_children():
            values = my_treeview.item(child)["values"]
            if len(values) > column_index:  # Đảm bảo cột tồn tại trong dữ liệu
                unique_values.add(values[column_index])
        return list(unique_values)

class utils_controller_validate_number_of_slip_is_exist_in_database:
    def start_validate(data=None, column_index=1, so_ky_tu=6, ma_tv=None, loai_phieu=None, database_name=None, table_name=None, column_name=None):
        new_data = utils_controller_validate_number_of_slip_is_exist_in_database.get_last_four_from_third_column(data, column_index, so_ky_tu)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text
        
        # Get the last two digits of the year
        year_last_two_digits = datetime.now().strftime("%y")
        
        # Dùng list comprehension để thêm tiền tố
        new_data = tuple(f'{ma_tv}-{loai_phieu}-{year_last_two_digits}{item}' for item in new_data)

        # Check exist
        flag, notification_text = utils_controller_check_exist.check_exist_values(new_data, database_name, table_name, column_name)
        if flag == False:
            return False, notification_text
        else:
            return True, "Số phiếu đã tồn tại."
        
    def get_last_four_from_third_column(data, column_index, so_ky_tu):
        # Lấy các giá trị duy nhất từ các ký tự cuối cùng
        unique_values = set(str(item[column_index])[-so_ky_tu:] for item in data)
        # Chuyển lại thành danh sách (nếu cần) và trả về
        return list(unique_values)
    
class utils_controller_validate_id_is_exist_in_database:
    def start_validate(data=None, column_index=1, database_name=None, table_name=None, column_name=None):
        new_data = utils_controller_validate_id_is_exist_in_database.get_unique_values_from_column(data, column_index)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        # Check exist
        flag, notification_text = utils_controller_check_exist.check_exist_values(new_data, database_name, table_name, column_name)
        if flag == False:
            return False, notification_text
        else:
            return True, "Các giá trị ID đã tồn tại."
        
    def get_unique_values_from_column(data, column_index):
        # Lấy các giá trị duy nhất
        unique_values = set(str(item[column_index]) for item in data)
        # Chuyển lại thành danh sách (nếu cần) và trả về
        return list(unique_values)

class utils_controller_validate_is_NULL:
    def start_validate(data=None, column_index=1):
        new_data = utils_controller_validate_is_NULL.get_values_from_column(data, column_index)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return True, notification_text

        # Kiểm tra các giá trị có phải là NULL
        for value in new_data:
            if value is None or (isinstance(value, float) and math.isnan(value)):
                notification_text = f"Có giá trị NULL hoặc NaN trong cột {column_index}"
                return True, notification_text
        
        notification_text = "Các giá trị không phải là NULL"
        return False, notification_text
        
    def get_values_from_column(data, column_index):
        values = [item[column_index] for item in data]
        # Chuyển lại thành danh sách (nếu cần) và trả về
        return list(values)


class utils_controller_validate_is_number:
    def start_validate(data=None, column_index=1):
        new_data = utils_controller_validate_is_NULL.get_values_from_column(data, column_index)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        # Kiểm tra các giá trị có phải là NULL
        for value in new_data:
            if not isinstance(value, (int, float)):
                notification_text = f"Có giá trị không phải số trong cột {column_index}"
                return False, notification_text
        
        notification_text = "Các giá trị không phải là NULL"
        return True, notification_text
        
    def get_values_from_column(data, column_index):
        values = [item[column_index] for item in data]
        # Chuyển lại thành danh sách (nếu cần) và trả về
        return list(values)
    
class utils_controller_validate_is_True_False:
    def start_validate(data=None, column_index=1):
        new_data = utils_controller_validate_is_NULL.get_values_from_column(data, column_index)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        # Kiểm tra các giá trị có phải là NULL
        for value in new_data:
            if value not in [True, False, 0, 1]:
                notification_text = f"Có giá trị không phải True hoặc False trong cột {column_index}"
                return False, notification_text
        
        notification_text = "Các giá trị không phải là True hoặc False"
        return True, notification_text
        
    def get_values_from_column(data, column_index):
        values = [item[column_index] for item in data]
        # Chuyển lại thành danh sách (nếu cần) và trả về
        return list(values)
        
class utils_controller_validate_logic_in_columns_num_01_nhu_cau_bang_tong_giucho_ycdh:
    def start_validate(data=None, column_indices=[1, 2, 3]):
        new_data = utils_controller_validate_logic_in_columns_num_01_nhu_cau_bang_tong_giucho_ycdh.get_values_from_columns(data, column_indices)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        for row in new_data:
            if len(row) < len(column_indices):
                notification_text = f"Dữ liệu không đủ {len(column_indices)} cột để kiểm tra."
                return True, notification_text
            
            last_column_index = len(column_indices)
            # Kiểm tra các giá trị có phải là NULL
            if row[0] != sum(row[1:last_column_index]):
                notification_text = f"Cột đầu không bằng tổng các cột sau {column_indices}"
                return False, notification_text
        
        notification_text = "Good logic in columns"
        return True, notification_text
        
    def get_values_from_columns(data, column_indices):
        values = [[row[i] for i in column_indices] for row in data]
        return values
        
class utils_controller_validate_logic_in_columns_num_02_one_slip_one_id:
    def start_validate(data=None, column_indices=[1, 2]):
        new_data = utils_controller_validate_logic_in_columns_num_02_one_slip_one_id.get_values_from_columns(data, column_indices)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        # Dùng dictionary để kiểm tra mỗi số phiếu có bao nhiêu mã khách hàng
        ticket_dict = {}

        for row in new_data:
            if len(row) < len(column_indices):
                return False, f"Dữ liệu không đủ {len(column_indices)} cột để kiểm tra."

            ticket, customer = row  # Lấy số phiếu và mã khách hàng

            if ticket not in ticket_dict:
                ticket_dict[ticket] = set()  # Tạo một tập hợp để lưu mã khách hàng

            ticket_dict[ticket].add(customer)

        # Kiểm tra xem có số phiếu nào có nhiều hơn 1 mã khách hàng không
        invalid_tickets = {ticket: customers for ticket, customers in ticket_dict.items() if len(customers) > 1}

        if invalid_tickets:
            error_messages = [
                f"Số phiếu {ticket} có nhiều mã khách hàng: {', '.join(customers)}"
                for ticket, customers in invalid_tickets.items()
            ]
            return False, "\n".join(error_messages)
        
        notification_text = "Good logic in columns"
        return True, notification_text
        
    def get_values_from_columns(data, column_indices):
        values = [[row[i] for i in column_indices] for row in data]
        return values
    
class utils_controller_validate_logic_in_columns_num_03_no_duplicate_inventory_id:
    def start_validate(data=None, column_indices=[1, 2]):
        new_data = utils_controller_validate_logic_in_columns_num_02_one_slip_one_id.get_values_from_columns(data, column_indices)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        # Dùng dictionary để kiểm tra mỗi số phiếu có bao nhiêu mã khách hàng
        ticket_item_dict = {}  # Dictionary lưu danh sách mã hàng theo số phiếu

        for row in new_data:
            if len(row) < len(column_indices):
                return False, f"Dữ liệu không đủ {len(column_indices)} cột để kiểm tra."

            ticket, item = row  # Lấy số phiếu và mã hàng

            # Kiểm tra một số phiếu không có hai dòng trùng mã hàng
            if ticket not in ticket_item_dict:
                ticket_item_dict[ticket] = set()
            
            if item in ticket_item_dict[ticket]:
                return False, f"Số phiếu {ticket} chứa mã hàng trùng nhau: {item}."
            
            ticket_item_dict[ticket].add(item)

        return True, "Dữ liệu hợp lệ: Không có số phiếu nào chứa mã hàng trùng nhau."
        
    def get_values_from_columns(data, column_indices):
        values = [[row[i] for i in column_indices] for row in data]
        return values

class utils_controller_validate_logic_in_columns_num_04_no_duplicate_stt_dong:
    """
    Cột thứ 2 (chỉ mục 1, tính từ 0) phải là duy nhất trong từng số phiếu.
    Cột thứ 2 phải là số nguyên.
    Cột thứ 2 phải bắt đầu từ số 1 và tiếp tục tăng liên tiếp.
    """
    
    def start_validate(data=None, column_indices=[1, 2]):
        new_data = utils_controller_validate_logic_in_columns_num_03_no_duplicate_inventory_id.get_values_from_columns(data, column_indices)
        if not new_data:
            notification_text = "Không có data để kiểm tra."
            return False, notification_text

        ticket_item_dict = {}  # Dictionary lưu danh sách mã hàng theo số phiếu
        expected_value_dict = {}  # Dictionary để lưu giá trị mong đợi theo từng số phiếu

        for row in new_data:
            if len(row) < len(column_indices):
                return False, f"Dữ liệu không đủ {len(column_indices)} cột để kiểm tra."

            ticket, item = row  # Lấy số phiếu và mã hàng

            # Kiểm tra nếu giá trị item là None hoặc NaN
            if item is None or (isinstance(item, float) and math.isnan(item)):
                return False, f"Có giá trị None hoặc NaN ở cột thứ 2, vui lòng kiểm tra lại dữ liệu."

            # Chuyển đổi số thực thành số nguyên nếu nó là số nguyên (VD: 16.0 -> 16)
            if isinstance(item, float) and item.is_integer():
                item = int(item)
            
            # Kiểm tra cột thứ 2 là số nguyên dương
            if not isinstance(item, int) or item <= 0:
                return False, f"Giá trị ở cột thứ 2 phải là số nguyên dương: {item}."
            
            # Kiểm tra cột thứ 2 là số nguyên
            try:
                item_int = int(item)
            except ValueError:
                return False, f"Giá trị ở cột thứ 2 không phải là số nguyên: {item}."
            
            # Kiểm tra số phiếu đã xuất hiện hay chưa
            if ticket not in ticket_item_dict:
                ticket_item_dict[ticket] = set()
                expected_value_dict[ticket] = 1  # Giá trị mong đợi đầu tiên cho số phiếu này
            
            # Kiểm tra số phiếu không có hai dòng trùng mã hàng
            if item_int in ticket_item_dict[ticket]:
                return False, f"Số phiếu {ticket} chứa mã hàng trùng nhau: {item_int}."
            
            ticket_item_dict[ticket].add(item_int)
            
            # Kiểm tra cột thứ 2 bắt đầu từ 1 và liên tiếp trong từng số phiếu
            if item_int != expected_value_dict[ticket]:
                return False, f"Số phiếu {ticket} có STT dòng không theo thứ tự liên tiếp, mong đợi {expected_value_dict[ticket]}, nhận {item_int}."
            expected_value_dict[ticket] += 1

        return True, "Dữ liệu hợp lệ: Mỗi số phiếu có giá trị duy nhất, là số nguyên và bắt đầu từ 1 theo thứ tự liên tiếp."
        
    def get_values_from_columns(data, column_indices):
        values = [[row[i] if row[i] is not None and not (isinstance(row[i], float) and math.isnan(row[i])) else None for i in column_indices] for row in data]
        return values

class utils_controller_get_data_to_print:
    def get_treeview_data(treeview: ttk.Treeview, columns: tuple) -> list:
        """
        Trích xuất dữ liệu từ Treeview.
        
        Args:
            treeview (ttk.Treeview): Treeview cần lấy dữ liệu.
            columns (tuple): Tuple chứa index các cột cần lấy dữ liệu.

        Returns:
            list: Danh sách các tuple, mỗi tuple chứa dữ liệu của một hàng theo các cột được chọn.
        """
        data = []
        for item in treeview.get_children():
            row = tuple(treeview.item(item, 'values')[col] for col in columns)
            data.append(row)
        return data
    
class utils_controller_Export_data_to_print_template:
    def start_loading_data_to_print_template(data_information_of_slip, data_information_from_treeview, template_path):
        try:
            # Đếm số dòng trên treeview
            
            # Xác định vị trí bắt đầu information of slip
            
            # Xác định vị trí bắt đầu information from treeview
            
            # Thêm dòng trên template
            
            # load data into template: info of slip
            
            # load data into template: info of treeview
            
            # Load data to template
            flag = utils_controller_Export_data_to_print_template.load_data_to_print_template(data_information_of_slip, data_information_from_treeview, template_path)
            if flag == False:
                return False
            else:
                return True
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return False
        
    def start_loading_data_to_print_template(data_information_of_slip, data_information_from_treeview, template_path):
        print(data_information_of_slip)