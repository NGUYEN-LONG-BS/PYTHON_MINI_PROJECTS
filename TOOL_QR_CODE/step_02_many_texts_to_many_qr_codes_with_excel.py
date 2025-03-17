import tkinter as tk
from tkinter import filedialog
import qrcode
from PIL import ImageTk, Image
import openpyxl
import os

# Hàm tạo QR code từ dữ liệu trong cột A và lưu vào cột B
def generate_qr_from_excel():
    # Chọn file Excel
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if not file_path:
        return  # Nếu không chọn file, dừng lại

    # Mở file Excel và sheet đầu tiên
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Đọc dữ liệu từ cột A (từ A1 đến A cuối cùng có dữ liệu)
    max_row = sheet.max_row
    data = [sheet.cell(row=i, column=1).value for i in range(1, max_row+1)]
    
    # Tạo mã QR và ghi vào cột B
    for i, text in enumerate(data, start=1):
        if text:  # Nếu ô không rỗng
            # Tạo QR code
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(text)
            qr.make(fit=True)
            
            # Tạo hình ảnh QR code
            img = qr.make_image(fill="black", back_color="white")
            
            # Lưu QR code vào cột B
            img_path = f"qr_{i}.png"
            img.save(img_path)
            
            # Lưu ảnh vào ô B
            img_file = open(img_path, 'rb')
            img_bytes = img_file.read()
            img_file.close()
            
            # Chèn ảnh vào cột B
            img = openpyxl.drawing.image.Image(img_path)
            img.width = 100  # Điều chỉnh kích thước QR code
            img.height = 100
            sheet.add_image(img, f'B{i}')
            
            # Điều chỉnh chiều cao dòng để vừa với kích thước của QR code
            sheet.row_dimensions[i].height = 100  # Chiều cao dòng 100 (tương đương với chiều cao của QR code)

            # Đặt lại kích thước cột B theo chiều rộng của QR code
            sheet.column_dimensions['B'].width = 15  # Có thể điều chỉnh theo yêu cầu

    # Chọn nơi lưu file kết quả
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="file_result")
    if save_path:
        # Nếu file đã tồn tại, tự động thêm số vào tên file
        base_name, ext = os.path.splitext(save_path)
        counter = 1
        while os.path.exists(save_path):
            save_path = f"{base_name}_{counter}{ext}"
            counter += 1

        # Lưu file kết quả
        wb.save(save_path)
        print("File đã được lưu tại:", save_path)

# Tạo cửa sổ chính
root = tk.Tk()
root.title("Tạo QR Code từ Excel")

# Tạo nút để chọn file Excel và tạo QR Code
generate_button = tk.Button(root, text="Chọn File Excel và Tạo QR Code", command=generate_qr_from_excel)
generate_button.pack(pady=20)

# Khởi chạy giao diện
root.mainloop()
