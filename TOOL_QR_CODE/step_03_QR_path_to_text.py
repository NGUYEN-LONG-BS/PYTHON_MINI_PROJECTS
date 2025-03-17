import tkinter as tk
from tkinter import filedialog
import openpyxl
from pyzbar.pyzbar import decode
from PIL import Image
import os

# Hàm giải mã QR code từ cột B và ghi kết quả vào cột C
def decode_qr_from_excel():
    # Chọn file Excel
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if not file_path:
        return  # Nếu không chọn file, dừng lại

    # Mở file Excel và sheet đầu tiên
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Đọc dữ liệu từ cột B (từ B1 đến B cuối cùng có dữ liệu)
    max_row = sheet.max_row
    for i in range(1, max_row + 1):
        qr_image_path = sheet.cell(row=i, column=2).value  # Đọc đường dẫn hình ảnh QR từ cột B
        if qr_image_path and os.path.exists(qr_image_path):  # Kiểm tra nếu đường dẫn hình ảnh tồn tại
            try:
                # Giải mã QR code
                img = Image.open(qr_image_path)
                decoded_objects = decode(img)
                if decoded_objects:
                    qr_text = decoded_objects[0].data.decode('utf-8')  # Lấy dữ liệu từ QR code
                    # Ghi kết quả vào cột C
                    sheet.cell(row=i, column=3, value=qr_text)
                else:
                    sheet.cell(row=i, column=3, value="Không đọc được QR code")
            except Exception as e:
                sheet.cell(row=i, column=3, value=f"Lỗi: {str(e)}")
    
    # Chọn nơi lưu file kết quả
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="file_result_decoded")
    if save_path:
        # Nếu file đã tồn tại, tự động thêm số vào tên file
        base_name, ext = os.path.splitext(save_path)
        counter = 1
        while os.path.exists(save_path):
            save_path = f"{base_name}_{counter}{ext}"
            counter += 1

        # Lưu file kết quả
        wb.save(save_path)
        print("File đã được lưu tại:", save_path)

# Tạo cửa sổ chính
root = tk.Tk()
root.title("Giải mã QR Code từ Excel")

# Tạo nút để chọn file Excel và giải mã QR code
generate_button = tk.Button(root, text="Chọn File Excel và Giải mã QR Code", command=decode_qr_from_excel)
generate_button.pack(pady=20)

# Khởi chạy giao diện
root.mainloop()
