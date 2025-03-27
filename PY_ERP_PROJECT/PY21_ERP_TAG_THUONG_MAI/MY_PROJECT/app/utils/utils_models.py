import os
import json
import pyodbc
from datetime import datetime
from openpyxl import load_workbook
import xlwings as xw

# Import từ chính thư mục utils
from . import utils_functions
from . import utils_controllers

class utils_model_TreeviewConfigLoader_250217_13h20:
    """Class để load cấu hình Treeview từ JSON"""

    def load_json(config_json_path):
        """Đọc dữ liệu từ file JSON"""
        try:
            with open(config_json_path, 'r', encoding='utf-8') as file:
                return json.load(file)
        except FileNotFoundError:
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            print(f"Error: File '{config_json_path}' không tồn tại.")
        except json.JSONDecodeError:
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            print("Error: JSON bị lỗi hoặc không hợp lệ.")
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
        return None

    def get_column_config(data):
        """Lấy cấu hình cột từ JSON"""
        try:
            columns = data["table"]["columns"]
            column_names = [col["name"] for col in columns]
            return columns, column_names
        except KeyError:
            print("Error: Thiếu key 'table' hoặc 'columns' trong JSON.")
            return [], []

    def get_header_font(data):
        """Lấy cấu hình font của tiêu đề"""
        try:
            font_config = data["table"]["headings"]
            return (font_config["family"], font_config["size"], font_config["weight"])
        except KeyError:
            print("Warning: Không tìm thấy cấu hình font. Dùng mặc định.")
            return ("Arial", 12, "normal")

    def extract_column_attribute(data, attribute, default_value):
        """Trích xuất thuộc tính của cột"""
        try:
            columns = data["table"]["columns"]
            return [column.get(attribute, default_value) for column in columns]
        except KeyError:
            print(f"Warning: Không tìm thấy key '{attribute}' trong JSON.")
            return []

class utils_model_get_data_from_SQL:
    
    def get_data_with_query(query):
        try:
            data = utils_model_get_data_from_SQL.fetch_data(query)
            if not data:
                return []
            return data
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return []
    
    def fetch_data(query):
        try:
            data = utils_functions.f_utils_fetch_data_from_database(query)
            if data:
                return data
            else:
                return []
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return []

class utils_model_get_data_from_Excel_250221_16h45:
    # Hàm để mở file Excel và lấy dữ liệu từ sheet, cho phép truyền vào ô bắt đầu
    def get_data_from_excel_with_xlwings(file_path, sheet_name, start_row=1, start_col=1):
        try:
            # Tạo ứng dụng Excel mà không hiển thị giao diện
            app = xw.App(visible=False)
            
            # Mở workbook từ file
            wb = app.books.open(file_path)
            
            # Lấy sheet cần làm việc
            sheet = wb.sheets[sheet_name]
            
            # Lấy dữ liệu từ ô bắt đầu được chỉ định
            # Expand phạm vi để lấy hết bảng dữ liệu
            data = sheet.range((start_row, start_col)).current_region.value
            # print("data with xlwings: ", data)
            
            # Đóng workbook và thoát ứng dụng Excel
            wb.close()
            app.quit()
            
            # Đảm bảo rằng đối tượng app đã được giải phóng
            del wb
            del app
            
            return data

        except Exception as e:
            print(f"Error: {e}")
            return f"Error: {e}"
    
    # Sử dụng openpyxl để đọc file Excel (không cần cài Excel)
    def get_data_from_excel_with_openpyxl(file_path, sheet_name, start_row=1, start_col=1):
        # Mở file Excel
        wb = load_workbook(file_path, data_only=True)   # Đảm bảo rằng ta đọc giá trị tính toán
        # print(wb)

        sheet = wb[sheet_name]
        
        # Lấy dữ liệu từ ô bắt đầu được chỉ định
        data = []
        for row in sheet.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
            data.append(row)
        # print("data with openpyxl: ", data)
        
        # Đóng workbook sau khi hoàn thành
        wb.close()
        # print(wb)
        
        # Đảm bảo rằng đối tượng app đã được giải phóng
        del wb
            
        # Trả về dữ liệu
        return data
    
class utils_model_import_data_to_SQL_SERVER_250221_16h45:
    def f_insert_data_to_sql(entry_notification, server_name, database_name, login_name, login_pass, table_name, data_array):
        """
        Hàm kết nối SQL Server và chèn dữ liệu từ mảng vào bảng, bỏ qua các cột có giá trị mặc định (ID, NGAY_TAO_PHIEU).
        
        :param server_name: Tên hoặc địa chỉ IP của máy chủ SQL Server.
        :param database_name: Tên cơ sở dữ liệu.
        :param login_name: Tên người dùng SQL Server.
        :param login_pass: Mật khẩu SQL Server.
        :param table_name: Tên bảng trong cơ sở dữ liệu.
        :param data_array: Mảng chứa dữ liệu cần chèn (list of lists).
        """
        # Kết nối đến SQL Server
        connection_string = f"DRIVER={{SQL Server}};SERVER={server_name};DATABASE={database_name};UID={login_name};PWD={login_pass}"
        try:
            conn = pyodbc.connect(connection_string)
            # print("Kết nối thành công đến cơ sở dữ liệu.")
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            utils_controllers.utils_controller_config_notification_250220_10h05.f_config_notification(entry_notification, "Kết nối cơ sở dữ liệu không thành công", "red")
            return False

        cursor = conn.cursor()
        
        # Lấy danh sách cột của bảng
        try:
            cursor.execute(f"SELECT * FROM {table_name} WHERE 1=0")
            columns = [column[0] for column in cursor.description]  # Lấy tên cột
            # print("Danh sách cột trong bảng:", columns)
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return False

        # Loại bỏ các cột có giá trị mặc định (ID, NGAY_TAO_PHIEU)
        columns_to_insert = [col for col in columns if col not in ['ID', 'DATE']]
        # print("Danh sách cột cần chèn:", columns_to_insert)
        
        # Kiểm tra số cột trong dữ liệu khớp với số cột cần chèn
        num_columns_to_insert = len(columns_to_insert)
        # print("num_columns_to_insert", num_columns_to_insert)

        # Kiểm tra dữ liệu xem có đúng số cột cần chèn không
        for row in data_array:
            if len(row) != num_columns_to_insert:
                utils_controllers.utils_controller_config_notification_250220_10h05.f_config_notification(entry_notification, f"Dữ liệu không khớp số cột cần chèn. Row: {row}", "red")
                return False

        # Chèn dữ liệu
        # print("Bắt đàu chèn dữ liệu...")
        try:
            placeholders = ", ".join(["?" for _ in range(num_columns_to_insert)])  # Tạo chuỗi placeholder "?, ?, ?"
            query = f"INSERT INTO {table_name} ({', '.join(columns_to_insert)}) VALUES ({placeholders})"
            
            # print("Placeholders: ", placeholders)
            # print("Query: ", query)
            
            for row in data_array:
                # print("Execute... ", row)
                cursor.execute(query, row)
                # print("Executed")
            
            # print("Commit...")
            conn.commit()
            # print("Dữ liệu đã được chèn thành công.")
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return False
        finally:
            cursor.close()
            conn.close()
            # print("Kết nối đã được đóng.")
            return True


class utils_model_get_data_from_SQL:
    def get_data_with_query(query):
        data = utils_model_SQL_server.fetch_data(query)
        return data

class utils_model_SQL_server:

    def fetch_data(query):
        try:
            data = utils_functions.f_utils_fetch_data_from_database(query)
            # print(data)
            return data
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return []
        
    def fetch_data_with_quey_and_params(query, params_list):
        try:
            data = utils_functions.f_utils_fetch_data_from_database_with_quey_and_params(query, params_list)
            # print(data)
            return data
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return []    
        
    def sent_SQL_query(query):
        try:
            utils_functions.f_utils_sent_query_to_SQL(query)
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
    
    def f_02_insert_data_to_sql(server_name, database_name, login_name, login_pass, table_name, data_array):
        """
        Hàm kết nối SQL Server và chèn dữ liệu từ mảng vào bảng, bỏ qua các cột có giá trị mặc định (ID, NGAY_TAO_PHIEU).
        
        :param server_name: Tên hoặc địa chỉ IP của máy chủ SQL Server.
        :param database_name: Tên cơ sở dữ liệu.
        :param login_name: Tên người dùng SQL Server.
        :param login_pass: Mật khẩu SQL Server.
        :param table_name: Tên bảng trong cơ sở dữ liệu.
        :param data_array: Mảng chứa dữ liệu cần chèn (list of lists).
        """
        # Kết nối đến SQL Server
        connection_string = f"DRIVER={{SQL Server}};SERVER={server_name};DATABASE={database_name};UID={login_name};PWD={login_pass}"
        try:
            conn = pyodbc.connect(connection_string)
            # print("Kết nối thành công đến cơ sở dữ liệu.")
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return

        cursor = conn.cursor()
        
        # Lấy danh sách cột của bảng
        try:
            cursor.execute(f"SELECT * FROM {table_name} WHERE 1=0")
            columns = [column[0] for column in cursor.description]  # Lấy tên cột
            # print("Danh sách cột trong bảng:", columns)
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return

        # Loại bỏ các cột có giá trị mặc định (ID, NGAY_TAO_PHIEU)
        columns_to_insert = [col for col in columns if col not in ['ID', 'DATE']]
        # print("Danh sách cột cần chèn:", columns_to_insert)
        
        # Kiểm tra số cột trong dữ liệu khớp với số cột cần chèn
        num_columns_to_insert = len(columns_to_insert)
        # print("num_columns_to_insert", num_columns_to_insert)
        if not all(len(row) == num_columns_to_insert for row in data_array):
            print("Dữ liệu không khớp số cột cần chèn.")
            return

        # Chèn dữ liệu
        try:
            placeholders = ", ".join(["?" for _ in range(num_columns_to_insert)])  # Tạo chuỗi placeholder "?, ?, ?"
            query = f"INSERT INTO {table_name} ({', '.join(columns_to_insert)}) VALUES ({placeholders})"
            
            for row in data_array:
                cursor.execute(query, row)
            
            conn.commit()
            # print("Dữ liệu đã được chèn thành công.")
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
        finally:
            cursor.close()
            conn.close()
            # print("Kết nối đã được đóng.")
    
    def f_goi_ham_Export_to_table(data_array, table_name):
        try:
            server_name = utils_functions.f_utils_get_DB_HOST()
            login_name, login_pass = utils_functions.f_utils_get_DB_USER_AND_DB_PASSWORD()
            database_name = utils_controllers.utils_controller_get_information_of_database.load_database_name()

            utils_model_SQL_server.f_02_insert_data_to_sql(server_name, database_name, login_name, login_pass, table_name, data_array)
            return True
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return False
    
    def f_validate_data_format_KD02_YEU_CAU_DAT_HANG(data_array):
        """
        Validate the format of data before inserting into SQL Server.
        :param data_array: List of tuples containing the data to validate.
        :return: True if all data is valid, otherwise False with error messages.
        """
        is_valid = True
        for idx, row in enumerate(data_array):
            try:
                # Validate each field
                column = 0
                if not isinstance(row[column], str) or len(row[column]) > 10:
                    raise ValueError(f"Data validation: ID_NHAN_VIEN (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 10.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: XOA_SUA (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                try:
                    datetime.strptime(row[column], '%Y-%m-%d')
                except ValueError:
                    raise ValueError(f"Data validation: NGAY_TREN_PHIEU (Row {idx+1}, Value: {row[column]}) must be a valid date in 'YYYY-MM-DD' format.")
                column = column + 1
                if not isinstance(row[column], str) or len(row[column]) > 50:
                    raise ValueError(f"Data validation: SO_PHIEU (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 50.")
                column = column + 1
                if not isinstance(row[column], str) or len(row[column]) > 50:
                    raise ValueError(f"Data validation: MA_DOI_TUONG (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 50.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: TEN_DOI_TUONG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: MST (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: DIA_CHI (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: SO_HOP_DONG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: THONG_TIN_HOP_DONG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: GHI_CHU_PHIEU (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], int) or row[column] < 1:
                    raise ValueError(f"Data validation: STT_DONG (Row {idx+1}, Value: {row[column]}) must be an integer greater than 0.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: MA_HANG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: TEN_HANG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: DVT (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] < 0:
                    raise ValueError(f"Data validation: SO_LUONG_KHA_DUNG (Row {idx+1}, Value: {row[column]}) must be a positive number.")
                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] <= 0:
                    raise ValueError(f"Data validation: SO_LUONG_NHU_CAU (Row {idx+1}, Value: {row[column]}) must be a positive number.")
                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] < 0:
                    raise ValueError(f"Data validation: SO_LUONG_GIU_CHO (Row {idx+1}, Value: {row[column]}) must be a positive number.")
                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] < 0:
                    raise ValueError(f"Data validation: SO_LUONG_YCDH (Row {idx+1}, Value: {row[column]}) must be a positive number.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: GHI_CHU_SP (Row {idx+1}, Value: {row[column]}) must be a string.")
            
            except ValueError as e:
                is_valid = False
                print(e)

        return is_valid
    
    def f_validate_data_format_KD02_KE_HOACH_DAT_HANG(data_array):
        """
        Validate the format of data before inserting into SQL Server.
        :param data_array: List of tuples containing the data to validate.
        :return: True if all data is valid, otherwise False with error messages.
        """
        is_valid = True
        for idx, row in enumerate(data_array):
            try:
                # Validate each field
                column = 0
                if not isinstance(row[column], str) or len(row[column]) > 10:
                    raise ValueError(f"Data validation: ID_NHAN_VIEN (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 10.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: XOA_SUA (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                try:
                    datetime.strptime(row[column], '%Y-%m-%d')
                except ValueError:
                    raise ValueError(f"Data validation: NGAY_TREN_PHIEU (Row {idx+1}, Value: {row[column]}) must be a valid date in 'YYYY-MM-DD' format.")
                column = column + 1
                if not isinstance(row[column], str) or len(row[column]) > 50:
                    raise ValueError(f"Data validation: SO_PHIEU (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 50.")
                column = column + 1
                if not isinstance(row[column], str) or len(row[column]) > 50:
                    raise ValueError(f"Data validation: MA_DOI_TUONG (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 50.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: TEN_DOI_TUONG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: MST (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: DIA_CHI (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: SO_HOP_DONG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: THONG_TIN_HOP_DONG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: GHI_CHU_PHIEU (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], int) or row[column] < 1:
                    raise ValueError(f"Data validation: STT_DONG (Row {idx+1}, Value: {row[column]}) must be an integer greater than 0.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: MA_HANG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: TEN_HANG (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: DVT (Row {idx+1}, Value: {row[column]}) must be a string.")
                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] <= 0:
                    raise ValueError(f"Data validation: SO_LUONG_NHU_CAU (Row {idx+1}, Value: {row[column]}) must be a positive number.")
                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: GHI_CHU_SP (Row {idx+1}, Value: {row[column]}) must be a string.")
            
            except ValueError as e:
                is_valid = False
                print(e)

        return is_valid
    
    def f_validate_data_format_TB_INVENTORY_CATEGORIES_250214_09h05(data_array):
        """
        Validate the format of data before inserting into SQL Server.
        :param data_array: List of tuples containing the data to validate.
        :return: True if all data is valid, otherwise False with error messages.
        """
        is_valid = True
        for idx, row in enumerate(data_array):
            try:
                # Validate each field
                column = 0
                if not isinstance(row[column], str) or len(row[column]) > 10:
                    raise ValueError(f"Data validation: ID_NHAN_VIEN (Row {idx+1}, Value: {row[column]}) must be a string with a maximum length of 10.")

                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: XOA_SUA (Row {idx+1}, Value: {row[column]}) must be a string.")

                column = column + 1
                if not isinstance(row[column], str) or row[column] is None:
                    raise ValueError(f"Data validation: MA_HANG (Row {idx+1}, Value: {row[column]}) must be a valid string and not null.")
                if len(row[column]) > 50:
                    raise ValueError(f"Data validation: MA_HANG (Row {idx+1}, Value: {row[column]}) must not exceed 50 characters.")

                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: TEN_HANG (Row {idx+1}, Value: {row[column]}) must be a string.")

                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: DVT (Row {idx+1}, Value: {row[column]}) must be a string.")

                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] < 0:
                    raise ValueError(f"Data validation: SL_TON_DAU_KY (Row {idx+1}, Value: {row[column]}) must be a positive number.")

                column = column + 1
                if not isinstance(row[column], (int, float)) or row[column] < 0:
                    raise ValueError(f"Data validation: DON_GIA_TON_DAU_KY (Row {idx+1}, Value: {row[column]}) must be a positive number.")

                column = column + 1
                if not isinstance(row[column], str):
                    raise ValueError(f"Data validation: MA_KHO_LUU_TRU (Row {idx+1}, Value: {row[column]}) must be a string.")

            
            except ValueError as e:
                is_valid = False
                print(e)

        return is_valid
    