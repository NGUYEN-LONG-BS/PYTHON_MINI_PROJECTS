
# Import từ chính thư mục utils
from . import utils_functions
from . import utils_models

import traceback
from datetime import datetime

import os
import sys
import time
import inspect
import tkinter as tk
from tkinter import ttk
from tkinter import font
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import xlwings as xw
from define import *
import pyodbc
import json
from cryptography.fernet import Fernet
import datetime
from datetime import datetime
from decimal import Decimal
from PIL import Image, ImageTk
import pandas as pd
import re

class utils_controller_action_after_event_250216_14h57:
    
    def f_get_the_latest_number_of_slip(entry_so_phieu, ma_thanh_vien, loai_phieu):
        # Get the latest number of slip
        so_phieu = utils_controller_get_the_latest_number_of_slip.handle_button_get_number_of_slip_click()
        
        # Create the connection string
        connection_number_of_slip = f"{ma_thanh_vien}-{loai_phieu}-{so_phieu + 1}"
        
        # Config the entry_so_phieu
        entry_so_phieu.config(state="normal")
        entry_so_phieu.delete(0, tk.END)
        entry_so_phieu.insert(0, connection_number_of_slip)
        entry_so_phieu.config(state="readonly")

    def update_entry_id_after_adding_new_row(tree, entry_id):
        row_count = 1 + len(tree.get_children())    
        entry_id.config(state="normal")  # Enable the Entry widget to update the value
        entry_id.delete(0, tk.END)  # Clear the existing value
        entry_id.insert(0, row_count)  # Insert the new value (ID)
        entry_id.config(state="disabled")  # Disable the Entry widget again

class utils_controller_config_notification_250220_10h05:
    def f_config_notification(element_label, text="...", fg="blue"):
            element_label.config(text=text, fg=fg)

class utils_controller_Export_data_to_Excel_250222_09h16:
    
    def export_log_to_excel(query, my_excel_header):
        # Truyền câu query vào model để lấy dữ liệu từ SQL (Giả sử hàm này trả về DataFrame)
        df = utils_models.utils_model_get_data_from_SQL.get_data_with_query(query)
        
        # Kiểm tra nếu có dữ liệu trả về
        if isinstance(df, list) and df:
            path = utils_controller_Export_data_to_Excel_250222_09h16.f_utils_export_data_to_excel(my_excel_header, df)
            if path:
                return True, path
        return False, None
    
    def f_utils_export_data_to_excel(data_header, data):
        # Kiểm tra tính nhất quán của dữ liệu
        if not data or not data_header:
            print("No data or headers provided.")
            return
        
        # Chuyển đổi dữ liệu từ pyodbc.Row thành tuple (nếu cần)
        data = [tuple(row) for row in data]

        # Kiểm tra số lượng cột có khớp không
        if len(data[0]) != len(data_header):
            print(f"Error: Mismatch in column count. Expected {len(data_header)}, but got {len(data[0])}.")
            return
        
        # Tạo DataFrame
        df = pd.DataFrame(data, columns=data_header)
        
        # Chuyển đổi kiểu dữ liệu Decimal và datetime để tương thích với Excel
        for column in df.columns:
            df[column] = df[column].apply(lambda x: float(x) if isinstance(x, Decimal) 
                                        else x.strftime('%Y-%m-%d %H:%M:%S') if isinstance(x, datetime) 
                                        else x)
        
        # Đường dẫn thư mục lưu file
        directory = PATH_DEFAUL
        if not os.path.exists(directory):
            os.makedirs(directory)  # Tạo thư mục nếu chưa tồn tại
        
        # Tạo file với tên không trùng
        file_path = utils_functions.f_utils_get_unique_filename(directory, "exported_data.xlsx")

        # Xuất ra file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Exported Data"
        
        # Ghi header
        ws.append(data_header)
        
        # Ghi dữ liệu
        for row in df.itertuples(index=False, name=None):
            ws.append(row)
        
        # Lưu file
        wb.save(file_path)
        wb.close()
        # print(f"Data exported successfully to {file_path}")
        utils_controller_Export_data_to_Excel_250222_09h16.format_excel_file(file_path, ws.title, "A1")
        return file_path
    
    def format_excel_file(file_path, sheet_name, start_cell):
        # Mở file Excel
        wb = openpyxl.load_workbook(file_path)
        
        # Kiểm tra nếu sheet tồn tại
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' không tồn tại trong {file_path}")
            return
        
        ws = wb[sheet_name]  # Chọn sheet theo tên
        
        # Xác định vị trí bắt đầu (ví dụ: "A1")
        start_row = ws[start_cell].row
        start_col = ws[start_cell].column

        # Tìm hàng và cột lớn nhất có dữ liệu
        max_row = ws.max_row
        max_col = ws.max_column

        # Xác định phạm vi dữ liệu (Dynamic)
        range_to_format = ws[start_cell:get_column_letter(max_col) + str(max_row)]

        # Áp dụng định dạng cho từng ô
        for row in range_to_format:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",  # Căn giữa ngang
                    vertical="center",    # Căn giữa dọc
                    wrap_text=False,      # Không xuống dòng tự động
                    shrink_to_fit=False   # Không co chữ vừa ô
                )

        # Tự động điều chỉnh độ rộng cột (AutoFit trong VBA)
        for col_num in range(start_col, max_col + 1):
            col_letter = get_column_letter(col_num)  # Lấy chữ cái của cột (A, B, C,...)
            max_length = 0

            for row_num in range(start_row, max_row + 1):
                cell_value = ws[f"{col_letter}{row_num}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws.column_dimensions[col_letter].width = max_length + 2  # Thêm padding

        # Lưu file
        wb.save(file_path)
        wb.close()


class utils_controller_get_the_latest_number_of_slip:
    
    def get_list_number_of_slip(database_name, table_name, slip_column_name):
        # Tạo câu query SQL với danh sách số phiếu
        query = f"""
        SELECT DISTINCT
            {slip_column_name}
        FROM [{database_name}].[dbo].[{table_name}]
        WHERE [XOA_SUA] = ''
        """
        # print("query", query)
        
        # lấy danh sách số phiếu từ SQL
        danh_sach_so_phieu = utils_models.utils_model_get_data_from_SQL.get_data_with_query(query)
        # print("danh_sach_so_phieu", danh_sach_so_phieu)
        
        return danh_sach_so_phieu
        
    def extract_numbers_from_data_SQL_num_01(data):
        data_01 = data
        data_02 = tuple(int(item[0].split('-')[-1]) for item in data_01)
        if not data_02:
            current_year = str(datetime.now().year)[-2:]  # Lấy 2 số cuối của năm hiện tại
            data_final = int(f'{current_year}0000')
        else:
            data_final = max(data_02)
        return data_final
        
    def handle_button_get_number_of_slip_click(database_name, table_name):
        # Lấy danh sách số phiếu từ SQL
        data_01 = utils_controller_get_the_latest_number_of_slip.get_list_number_of_slip(database_name, table_name)
        # Lấy số phiếu cuối cùng
        data_02 = utils_controller_get_the_latest_number_of_slip.extract_numbers_from_data_SQL_num_01(data_01)
        
        return data_02
   
class utils_controller_get_the_header_of_table_in_SQL_250221_11h01:
    def get_column_names(database_name, table_name):
        try:
            query = f"""
            SELECT COLUMN_NAME
            FROM [{database_name}].INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = '{table_name}' AND TABLE_SCHEMA = 'dbo'
            """
            # lấy danh sách tên cột từ SQL
            column_headers = utils_models.utils_model_get_data_from_SQL.get_data_with_query(query)
            if not column_headers:
                return None
            
            return column_headers
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", utils_functions.f_utils_get_current_function_name())
            return None
             
class utils_controller_set_size_of_windown_250215_10h24:    
    def f_utils_set_window_size_of_new_view(root, width=0, height=0, maximize=False):
        """Set the window size to 4/5 of the screen size or maximize it."""
        if maximize:
            root.state('zoomed')  # Maximizes the window (Windows)
            # root.is_maximized = True  # Set the flag to track maximized state
            # root.bind("<Configure>", utils_controller_set_size_of_windown_250215_10h24.on_resize)  # Bind the resize event
            return

        if width == 0 or height == 0:
            # Get screen dimensions
            screen_width = root.winfo_screenwidth()
            screen_height = root.winfo_screenheight()
            
            # Calculate 4/5 of screen dimensions
            ratio = 0.75
            height = int(screen_height * ratio)
            width = int(screen_width * ratio)
        
        # Set the window size
        root.geometry(f"{width}x{height}")

    # def on_resize(event):
    #     """Called when the window is resized, adjusts the window to 75% of the screen if maximized."""
    #     if isinstance(event.widget, tk.Tk):  # Ensure we're working with the root window
    #         root = event.widget
            
    #         # # Only apply the resizing logic if the window was maximized
    #         # if getattr(root, 'is_maximized', False):  # Check if window was maximized
    #         #     screen_width = root.winfo_screenwidth()
    #         #     screen_height = root.winfo_screenheight()

    #         #     # Calculate 75% of screen dimensions
    #         #     ratio = 0.75
    #         #     new_width = int(screen_width * ratio)
    #         #     new_height = int(screen_height * ratio)

    #         #     current_width = root.winfo_width()
    #         #     current_height = root.winfo_height()

    #         #     # Only update the window size if it's different from the current size
    #         #     if current_width != new_width or current_height != new_height:
    #         #         root.geometry(f"{new_width}x{new_height}")
    #         #         root.is_maximized = False
    #         # else:
    #         #     # If the window is not maximized, do nothing
    #         #     pass
            
    #         if root.winfo_width() < 1000 or root.winfo_height() < 750:
    #             root.geometry("1000x750")
            
class utils_controller_TreeviewConfigurator_250217_13h20:
    """Class để áp dụng cấu hình cho Treeview"""

    @staticmethod
    def apply_treeview_config(my_treeview, config_json_path):
        """Cấu hình Treeview dựa trên JSON"""
        
        # Xóa các cột hiện tại
        my_treeview.delete(*my_treeview.get_children())
        for col in my_treeview["columns"]:
            my_treeview.heading(col, text="")  # Xóa tiêu đề

        # Load dữ liệu từ JSON
        data = utils_models.utils_model_TreeviewConfigLoader_250217_13h20.load_json(config_json_path)
        if not data:
            return
        
        columns_config, column_names = utils_models.utils_model_TreeviewConfigLoader_250217_13h20.get_column_config(data)
        my_treeview["columns"] = column_names

        # Lấy cấu hình font header
        header_font = utils_models.utils_model_TreeviewConfigLoader_250217_13h20.get_header_font(data)

        # Cấu hình cột
        for config, col in zip(columns_config, column_names):
            my_treeview.heading(col, text=col)
            my_treeview.column(
                col,
                width=config.get("width", 100),
                minwidth=config.get("min_width", 50),
                anchor=config.get("anchor", "w"),
                stretch=config.get("stretch", True)
            )

        # Cấu hình font tiêu đề
        style = ttk.Style()
        style.configure("Treeview.Heading", font=header_font)

        # print("Treeview đã được cấu hình thành công từ JSON.")

class utils_controller_TreeviewHandler_click_250217_22h34:

    def treeview_single_click(my_Treeview):
        selected_item = my_Treeview.selection()
        if selected_item:
            row_values = my_Treeview.item(selected_item[0], "values")
            result_tuple = tuple(row_values) if row_values else None
            # print(result_tuple)
            return result_tuple
        return None

    def treeview_double_click(my_Treeview, column_return):
        selected_item = my_Treeview.selection()
        if selected_item:
            row_values = my_Treeview.item(selected_item[0], "values")
            return row_values[column_return] if len(row_values) > column_return else None
        return None
    
