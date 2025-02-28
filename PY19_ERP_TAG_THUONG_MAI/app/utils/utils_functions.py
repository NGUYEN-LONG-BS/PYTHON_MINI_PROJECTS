import os
import sys
import time
import inspect
import tkinter as tk
from tkinter import ttk
from tkinter import font
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import xlwings as xw
from define import *
import pyodbc
import json
from cryptography.fernet import Fernet
# import datetime
from datetime import datetime
from decimal import Decimal
from PIL import Image, ImageTk
import pandas as pd
import re

# Import từ chính thư mục utils
from . import utils_controllers
from . import utils_models

def f_utils_setup_logo(parent_frame):
    # Define function when click
    def on_logo_click(event):
        # Get the top-level window containing the parent_frame
        parent_window = parent_frame.winfo_toplevel()
        parent_window.destroy()  # Close the parent window
        # Open Dashboard
        from views.AD01_Dashboard_View.Dashboard_View import cls_Dashboard_View
        cls_Dashboard_View()
        
    try:
        # Try loading the light mode image first
        logo_image_light = Image.open(PATH_LOGO_LIGHT)
        logo_image_dark = Image.open(PATH_LOGO_DARK)
        
        # Convert the PIL image to a Tkinter-compatible photo image
        logo_image_light_tk = ImageTk.PhotoImage(logo_image_light)
        logo_image_dark_tk = ImageTk.PhotoImage(logo_image_dark)
        
        # Store the image as an attribute of the parent_frame (to prevent it from being garbage collected)
        parent_frame.logo_image_light = logo_image_light_tk
        parent_frame.logo_image_dark = logo_image_dark_tk
        
        # Create the tk.Label and display the image
        logo_label = tk.Label(parent_frame, 
                              image=logo_image_light_tk,
                              cursor="hand2")
        logo_label.pack(fill="both", expand=True) # Using pack to add it to the parent_frame
        logo_label.bind("<Button-1>", on_logo_click)  # Button-1 is left mouse click
    except FileNotFoundError:
        print(f"Logo file not found at {logo_image_light} or {logo_image_dark}")
        error_label = tk.Label(parent_frame, text="Logo not found", font=("", 16))
        error_label.pack(fill="both", expand=True)
        error_label.bind("<Button-1>", on_logo_click)  # Button-1 is left mouse click

def f_utils_setup_fav_icon(window):
        # Get the project root directory
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))  # Going 2 levels up
        
        # Load the logo images
        fav_icon = os.path.join(project_root, "assets/icons/favicon.png")
        
        img = Image.open(fav_icon)  # Replace with the path to your image file
        logo = ImageTk.PhotoImage(img)
        # Set the window icon
        window.iconphoto(False, logo)

def f_utils_find_my_function_path(function_name):
    source_file = inspect.getfile(function_name)
    print(f"Function is defined in: {source_file}")

def set_window_size(root, width=1600, height=900):
    # Thiết lập kích thước cửa sổ
    root.geometry(f"{width}x{height}")
    
    # Đảm bảo cửa sổ không thể thay đổi kích thước
    root.resizable(False, False)
    
    # Nếu bạn muốn căn giữa cửa sổ, bạn có thể tính toán vị trí và đặt lại
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int(screen_height / 2 - height / 2)
    position_right = int(screen_width / 2 - width / 2)
    print(f"Position Right: {position_right}, Position Top: {position_top}")
    
    root.geometry(f'{width}x{height}+{position_right}+{position_top}')

def f_utils_set_center_screen(root):
    # Lấy kích thước của cửa sổ
    root.update_idletasks()  # Cập nhật các thay đổi về kích thước
    width = root.winfo_width()
    height = root.winfo_height()
    
    # lấy thông tin kích thước màn hình và tinh toán lại
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Tính toán vị trí để căn giữa cửa sổ
    position_top = int(screen_height / 2 - height / 2) - 50
    position_right = int(screen_width / 2 - width / 2)

    # Đặt lại vị trí của cửa sổ
    root.geometry(f'{width}x{height}+{position_right}+{position_top}')

# Reusable function to set font for menu items
def f_utils_set_menu_font(widget, size=14, font_is="Arial"):
    """Set the font for menu items to a specified size."""
    custom_font = font.Font(family=font_is, size=size)  # Create a new Font object with the desired size
    widget.config(font=custom_font)  # Apply the font to the menu

def f_utils_open_dashboard_main():
    try:
        from views.AD01_Dashboard_View.Dashboard_View import cls_Dashboard_View
        new_view = cls_Dashboard_View()
        utils_controllers.utils_controller_set_size_of_windown_250215_10h24.f_utils_set_window_size_of_new_view(new_view, maximize=True)
        f_utils_set_center_screen(new_view)
        new_view.focus_force()
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
    
def f_utils_open_AI_chatbot():
    from views.AI00_MY_AI_ASSITANT.chat_view import AIChatApp
    new_window = tk.Toplevel()
    AIChatApp(new_window)
    
def f_utils_open_dashboard_kinh_doanh():
    try:
        from views.KD00_DashboardKinhDoanh_View.Dashboard_kinhdoanh_View import cls_Dashboard_kinhdoanh_View
        new_view = cls_Dashboard_kinhdoanh_View()
        utils_controllers.utils_controller_set_size_of_windown_250215_10h24.f_utils_set_window_size_of_new_view(new_view, maximize=True)
        f_utils_set_center_screen(new_view)
        new_view.focus_force()
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
    
def f_utils_open_dashboard_admin():
    try:
        from views.AD00_DashboardAdmin_View.Dashboard_admin_View import cls_Dashboard_admin_View
        new_view = cls_Dashboard_admin_View()
        utils_controllers.utils_controller_set_size_of_windown_250215_10h24.f_utils_set_window_size_of_new_view(new_view, maximize=True)
        f_utils_set_center_screen(new_view)
        new_view.focus_force()
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
    
def f_utils_open_dashboard_vat_tu():
    try:
        from views.VT00_DashboardVatTu_View.Dashboard_VatTu_View import cls_Dashboard_Vat_Tu_View
        new_view = cls_Dashboard_Vat_Tu_View()
        utils_controllers.utils_controller_set_size_of_windown_250215_10h24.f_utils_set_window_size_of_new_view(new_view, maximize=True)
        f_utils_set_center_screen(new_view)
        new_view.focus_force()
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())

def f_utils_show_fading_popup(message):
    # Tạo cửa sổ popup
    popup = tk.Toplevel()
    popup.title("Thông báo")
    # Set background color of popup window
    popup.config(bg=BG_COLOR_0_0)
    # Ẩn thanh tiêu đề (title bar)
    popup.overrideredirect(True)
    # Căn giữa màn hình
    utils_controllers.utils_controller_set_size_of_windown_250215_10h24.f_utils_set_window_size_of_new_view(popup, 150, 50, maximize=False)
    f_utils_set_center_screen(popup)
    
    # Add frame
    main_frame = tk.Frame(popup, width=popup.winfo_width(), height=popup.winfo_height(), bd=1, relief="solid")
    # main_frame.grid(row=0, column=0)
    main_frame.pack()

    # Tạo nhãn để hiển thị thông báo
    label = tk.Label(main_frame, text=message, font=("Helvetica", 12))
    label.pack(pady=10, padx=10)

    # Đặt thời gian để tự động đóng cửa sổ sau 3 giây (3000 milliseconds)
    popup.after(1000, popup.destroy)

def f_utils_tim_component_label_with_text(root=None, text_to_find=""):
    if root is None:  # Start from self if root is not provided
        # root = self
        return
    
    # Loop through all children of the current root
    for widget in root.winfo_children():
        # Check if widget is a Label and its text matches
        if isinstance(widget, tk.Label) and widget.cget("text") == text_to_find:
            return widget  # Found the Label, return it

        # If widget has children, recursively search in its children
        result = f_utils_tim_component_label_with_text(widget, text_to_find)
        if result:
            return result
    return None  # No matching Label found

def f_utils_tim_component_with_name(root=None, name_to_find=""):
    """
    Recursively searches for a Label widget with a specific name.
    :param root: The root widget to start the search from.
    :param name_to_find: The name of the widget to find.
    :return: The found Label widget, or None if not found.
    """
    if root is None:
        return None  # If no root is provided, return None

    # Iterate through all children of the current root
    for widget in root.winfo_children():
        # Check if the widget's name matches the target name
        if widget.winfo_name() == name_to_find:
            return widget  # Found the widget, return it

        # Recursively search in the widget's children
        result = f_utils_tim_component_with_name(widget, name_to_find)
        if result:
            return result

    return None  # No matching widget found

def f_utils_open_file():
    try:
        # Open file dialog to select a file
        file_path = filedialog.askopenfilename(title="Select a File")
        file_name = os.path.basename(file_path)
        if file_path:
            # Check if the file is an Excel file
            if file_path.endswith(('.xls', '.xlsx')):
                return file_name, file_path
            else:
                return "", ""
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return f"Error: {e}"

def f_utils_create_template_excel_file(file_name="template_wb.xlsx", sheet_name="template_sh", column_names=[('Col_01',), ('Col_02',), ('Col_03',)]):
    try:
        # Open a file dialog to select the save location
        export_path = PATH_DEFAUL
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialdir=export_path,
            initialfile=file_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File"
        )
        if not file_path:
            return False, ""
        
        # Create a new Excel workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        if os.path.exists(file_path):
            base, extension = os.path.splitext(file_path)
            counter = 1
            # Keep generating new filenames until an unused name is found
            while os.path.exists(file_path):
                file_path = f"{base} ({counter}){extension}"
                counter += 1
        
        # Write column names to the first row
        for col_num, column_name_tuple in enumerate(column_names, start=1):
            # Extract the string from the tuple
            column_name = column_name_tuple[0]
            cell = sheet.cell(row=1, column=col_num, value=column_name)
            # Apply formatting to the header row
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 15  # Adjust column width

        # Set background color for the header row
        for cell in sheet[1]:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
        # Auto-fit column widths based on content
        for col_num in range(1, len(column_names) + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)  # Add a little padding
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(file_path)
        
        # Close the workbook
        workbook.close()
        
        return True, f"Excel file created and saved to: {file_path}"
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return False, f"An error occurred: {str(e)}"

def f_utils_find_string_in_row_of_excel(file_path, sheet_name, target_string, row_number=1, case_sensitive=True, return_as_index=True):
    """
    Find a string in a specified row of an Excel sheet.
    
    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to search.
        target_string (str): String to search for.
        row_number (int): Row number to search in (default is 1).
        case_sensitive (bool): Whether the search should be case-sensitive (default is True).
        return_as_index (bool): Whether to return the column index (True) or column name (False).
    
    Returns:
        int or str: Column index or name where the string is found. Returns None if not found.
    """
    try:
        # Load the workbook and sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook[sheet_name]

        # Check for valid row number
        if row_number < 1 or row_number > sheet.max_row:
            raise ValueError(f"Invalid row number: {row_number}. It must be between 1 and {sheet.max_row}.")

        # Iterate through columns in the specified row
        for col in sheet.iter_cols(1, sheet.max_column):
            cell = col[row_number - 1]  # Row numbers are 1-based
            if cell.value is not None:
                cell_value = str(cell.value)
                if not case_sensitive:
                    if cell_value.lower() == target_string.lower():
                        return cell.column if return_as_index else cell.column_letter
                else:
                    if cell_value == target_string:
                        return cell.column if return_as_index else cell.column_letter
        
        return None  # String not found
    
    except FileNotFoundError:
        my_current_function = f_utils_get_current_function_name()
        print(f"Error at function: {my_current_function}")
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        raise Exception(f"An error occurred: {str(e)}")
        

def f_utils_find_string_in_column_of_excel(file_path, sheet_name, target_string, column_number=1, case_sensitive=True, return_as_index=True):
    """
    Find a string in a specified column of an Excel sheet.
    
    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to search.
        target_string (str): String to search for.
        column_number (int): Column number to search in (default is 1).
        case_sensitive (bool): Whether the search should be case-sensitive (default is True).
        return_as_index (bool): Whether to return the row index (True) or cell reference (False).
    
    Returns:
        int or str: Row index or cell reference where the string is found. Returns None if not found.
    """
    try:
        # Load the workbook and sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook[sheet_name]

        # Check for valid column number
        if column_number < 1 or column_number > sheet.max_column:
            raise ValueError(f"Invalid column number: {column_number}. It must be between 1 and {sheet.max_column}.")

        # Iterate through rows in the specified column
        for row in sheet.iter_rows(1, sheet.max_row):
            cell = row[column_number - 1]  # Column numbers are 1-based
            if cell.value is not None:
                cell_value = str(cell.value)
                if not case_sensitive:
                    if cell_value.lower() == target_string.lower():
                        return cell.row if return_as_index else f"{cell.column_letter}{cell.row}"
                else:
                    if cell_value == target_string:
                        return cell.row if return_as_index else f"{cell.column_letter}{cell.row}"
        
        return None  # String not found
    
    except FileNotFoundError:
        my_current_function = f_utils_get_current_function_name()
        print(f"Error at function: {my_current_function}")
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        raise Exception(f"An error occurred: {str(e)}")

def f_utils_copy_sheet_to_new_workbook(file_path, sheet_name):
    try:
        # Start a new Excel instance
        app = xw.App(visible=False, add_book=False)  # Start a new Excel process
        app.display_alerts = False  # Suppress Excel alerts
        # app.screen_updating = False  # Speed up operations

        # Open the workbook
        wb = app.books.open(file_path)

        # Check if the "PRINT" sheet exists
        if sheet_name in [sheet.name for sheet in wb.sheets]:
            print_sheet = wb.sheets[sheet_name]

            # Create a new workbook
            new_wb = xw.Book()

            # Copy the "PRINT" sheet to the new workbook
            print_sheet.api.Copy(Before=new_wb.sheets[0].api)
            print("Sheet {sheet_name} copied successfully.")
            
            # Make the new workbook visible to the user
            new_wb.app.visible = True
            print("New workbook opened for user.")

            # Maximize the new workbook window
            new_wb.app.api.WindowState = -4137  # -4137 corresponds to the "maximized" state
            print("New workbook maximized and opened for user.")
        else:
            print("Sheet 'PRINT' does not exist in the workbook.")

        # Close the original workbook
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
    finally:
        # Ensure the Excel application started by this code is closed
        if not app.books:  # If no books are left open, quit the app
            app.quit()
            
def f_utils_open_an_excel_with_path(file_path):
    try:
        # Mở file Excel bằng ứng dụng mặc định
        os.startfile(file_path)
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return f"Error: {e}"

def f_utils_delete_extend_row_and_column(file_path, sheet_name, start_column, end_column, start_row, last_row):
    """
    Delete rows and columns in an Excel sheet based on specified ranges.

    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to modify.
        start_column (int): Start column index to delete before.
        end_column (int): End column index to delete after.
        start_row (int): Start row index to delete before.
        last_row (int): Last row index to delete after.

    Returns:
        None: The function modifies the Excel file in place.
    """
    try:
        # Load the workbook and sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook[sheet_name]

        # Delete rows above start_row (if applicable)
        if start_row > 1:
            sheet.delete_rows(1, start_row - 1)

        # Delete rows below last_row (if applicable)
        if last_row < sheet.max_row:
            sheet.delete_rows(last_row + 1, sheet.max_row - last_row)

        # Delete columns to the left of start_column (if applicable)
        if start_column > 1:
            sheet.delete_cols(1, start_column - 1)

        # Delete columns to the right of end_column (if applicable)
        if end_column < sheet.max_column:
            sheet.delete_cols(end_column + 1, sheet.max_column - end_column)

        # Save the modified workbook
        workbook.save(file_path)
        print(f"Modifications completed successfully for '{sheet_name}' in file '{file_path}'.")

    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        raise Exception(f"An error occurred: {str(e)}")

def f_utils_copy_all_cells_and_paste_value(file_path, sheet_name):
    """
    Select all cells in the specified sheet and paste their values (removing any formulas).

    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to process.

    Returns:
        None: The function modifies the Excel file in place.
    """
    try:
        # Load the workbook and sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook[sheet_name]

        # Iterate through all cells in the sheet and copy values
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.value = cell.value  # Keep the value, remove any formulas

        # Save the modified workbook
        workbook.save(file_path)
        print(f"All cell values have been pasted as values in sheet '{sheet_name}'.")

    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        raise Exception(f"An error occurred: {str(e)}")

def f_utils_open_print_template(file_path, sheet_name):
    start_column = f_utils_find_string_in_row_of_excel(file_path, sheet_name, "FIRST_COLUMN", row_number=1, case_sensitive=True, return_as_index=True)
    end_column = f_utils_find_string_in_row_of_excel(file_path, sheet_name, "LAST_COLUMN", row_number=1, case_sensitive=True, return_as_index=True)
    value_column = f_utils_find_string_in_row_of_excel(file_path, sheet_name, "VALUE_COLUMN", row_number=1, case_sensitive=True, return_as_index=False)
    start_row = f_utils_find_string_in_column_of_excel(file_path, sheet_name, "FIRST_ROW", column_number=1, case_sensitive=True, return_as_index=True)
    last_row = f_utils_find_string_in_column_of_excel(file_path, sheet_name, "LAST_ROW", column_number=1, case_sensitive=True, return_as_index=True)
    
    print(start_column)
    print(end_column)
    print(value_column)
    print(start_row)
    print(last_row)
    
    f_utils_copy_sheet_to_new_workbook(file_path, sheet_name)
    # f_utils_delete_extend_row_and_column(file_path, sheet_name, start_column, end_column, start_row, last_row)
    
def f_utils_on_entry_change(entry_widget):
    """
    Callback to handle changes in the Entry widget.
    :param entry_widget: The Tkinter Entry widget being edited.
    """
    def format_entry_as_number(entry_widget):
        """
        Formats the content of the given Entry widget as a number with commas and two decimal places.
        :param entry_widget: The Tkinter Entry widget to format.
        """
        try:
            # Get the current text from the Entry widget
            current_text = entry_widget.get()

            # Remove any commas and validate as a number
            clean_text = current_text.replace(",", "")
                
            if clean_text.strip() == "":
                entry_widget.delete(0, tk.END)
                return

            if float(clean_text).is_integer():  # Nếu là số nguyên
                formatted_text = f"{int(clean_text):,}"
            else:  # Nếu là số thập phân
                formatted_text = f"{clean_text:,.2f}"

            # Update the Entry widget with the formatted text
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, formatted_text)
        except ValueError:
            # If the text is not a valid number, ignore the formatting
            pass
    entry_widget.after(100, lambda: format_entry_as_number(entry_widget))
    
def f_utils_access_widget_by_path(root, widget_path):
    try:
        widget = root.nametowidget(widget_path)
        # print(f"Found widget: {widget}")
        return widget
    except KeyError:
        print(f"Widget not found: {widget_path}")
        return None
    
def f_utils_set_style(object):
        # Create a style object
        style = ttk.Style()
        # Use default theme (you can experiment with others)
        # style.theme_use("default")
        style.theme_use("clam")

        # Customize the notebook style
        style.configure("TNotebook", 
                        background=BG_COLOR_0_0, 
                        borderwidth=0
                        )
        style.configure("TNotebook.Tab",
                        background=BG_COLOR_0_0,
                        foreground=FG_COLOR_01,
                        padding=[10, 5],
                        font=("Arial", 12, "bold"))
        style.map("TNotebook.Tab", 
                background=[("selected", HIGHLIGHT_COLOR)],
                foreground=[("selected", FG_COLOR_01)],
                padding=[("selected", [10, 5])],                                # Maintain padding without expanding borders
                expand=[("selected", [0, 0, 0, 0])])                            # No border expansion when selected
        
        # Customize the Label style
        style.configure("TLabel", 
                        background=BG_COLOR_0_0,  # Background color for labels
                        foreground=FG_COLOR_01, # Text color
                        font=("Arial", 10))     # Font style and size

        # Customize the Treeview style
        style.configure("Treeview", 
                        background=BG_COLOR_0_0, 
                        foreground=FG_COLOR_01, 
                        rowheight=25,          # Row height
                        fieldbackground=BG_COLOR_0_0, # Background color for the cells
                        font=("Arial", 10))
        style.map("Treeview", 
                background=[("selected", HIGHLIGHT_COLOR)],
                foreground=[("selected", FG_COLOR_01)])

        # Customize the Treeview heading style
        style.configure("Treeview.Heading", 
                        background=BG_COLOR_0_0, 
                        foreground=FG_COLOR_01, 
                        font=("Arial", 11, "bold")) # Font for headings

        # Customize the Scrollbar style
        style.configure("TScrollbar", 
                        background=BG_COLOR_0_0, 
                        troughcolor=BG_COLOR_0_2, # Trough (track) color
                        arrowcolor=FG_COLOR_03) # Arrow color
        style.map("TScrollbar", 
                background=[("pressed", HIGHLIGHT_COLOR), ("active", FG_COLOR_03)])
        
        # Customize the Frame style
        style.configure("TFrame", 
                        background=BG_COLOR_0_0,  # Background color for frames
                        borderwidth=2,           # Border width
                        relief="flat")           # Relief style (flat, raised, sunken, etc.)

        style.map("TFrame", 
                background=[("active", HIGHLIGHT_COLOR)])  # Optional: Change on hover or active state
        
def f_utils_get_unique_column_from_data(sample_data, number_column):
        # Trích xuất cột thứ n và loại bỏ trùng lặp
        unique_ma_hang = list(set(row[number_column] for row in sample_data))
        return unique_ma_hang

def f_utils_get_unique_column_from_treeview(treeview, number_column):
        # Extract unique values from the specified column in the Treeview
        unique_values = set()
        
        # Iterate through all items in the Treeview
        for item in treeview.get_children():
            value = treeview.item(item)['values'][number_column]
            unique_values.add(value)
        
        return list(unique_values)

def f_utils_get_formatted_today_YYYY_MM_DD(format_string='%Y-%m-%d'):
    # Get today's date
    today = datetime.today()
    # Format the date to 'YYYY-MM-DD'
    formatted_date = today.strftime(format_string)
    return formatted_date

def f_utils_change_format_date_from_ddmmyyyy_to_yyyymmdd(date_string):
    # Check if the input string is None or empty
    if not date_string:
        return ''  # Return empty string if input is null or empty
    
    # Use regex to extract day, month, and year regardless of separator (-, /, .)
    match = re.match(r"(\d{2})[\-/.](\d{2})[\-/.](\d{4})", date_string)
    
    if not match:
        return ''  # Return empty string if format is incorrect
    
    day, month, year = match.groups()
    
    # Reformat and return the date as yyyy-mm-dd
    return f"{year}-{month}-{day}"

def f_utils_get_unique_filename(directory, filename):
    """
    Kiểm tra xem file có tồn tại không. Nếu có, thêm số vào sau tên file để tránh trùng.
    Ví dụ: exported_data.xlsx → exported_data_1.xlsx → exported_data_2.xlsx
    """
    base, ext = os.path.splitext(filename)  # Tách tên file và phần mở rộng
    counter = 1
    new_filename = filename

    while os.path.exists(os.path.join(directory, new_filename)):
        new_filename = f"{base}_{counter}{ext}"
        counter += 1

    return os.path.join(directory, new_filename)

def f_utils_export_data_to_excel(data_header, data):
    # Kiểm tra tính nhất quán của dữ liệu
    if not data or not data_header:
        print("No data or headers provided.")
        return
    
    # Chuyển đổi dữ liệu từ pyodbc.Row thành tuple (nếu cần)
    data = [tuple(row) for row in data]

    # Kiểm tra số lượng cột có khớp không
    if len(data[0]) != len(data_header):
        print(f"Error: Mismatch in column count. Expected {len(data_header)}, but got {len(data[0])}.")
        return
    
    # Tạo DataFrame
    df = pd.DataFrame(data, columns=data_header)
    
    # Chuyển đổi kiểu dữ liệu Decimal và datetime để tương thích với Excel
    for column in df.columns:
        df[column] = df[column].apply(lambda x: float(x) if isinstance(x, Decimal) 
                                      else x.strftime('%Y-%m-%d %H:%M:%S') if isinstance(x, datetime) 
                                      else x)
    
    # Đường dẫn lưu file
    file_path = os.path.join(os.getcwd(), "exported_data.xlsx")

    # Xuất ra file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"
    
    # Ghi header
    ws.append(data_header)
    
    # Ghi dữ liệu
    for row in df.itertuples(index=False, name=None):
        ws.append(row)
    
    # Lưu file
    wb.save(file_path)
    print(f"Data exported successfully to {file_path}")

# ========================================================================================================================================================================
# Lấy thông tin từ file config.json
# ========================================================================================================================================================================
# Đọc encryption_key từ file
def load_encryption_key(file_path):
    with open(file_path, 'rb') as file:
        return file.read().strip()

# Function to load and decrypt configuration
def load_config(file_path, encryption_key):
    with open(file_path, 'r') as file:
        encrypted_config = json.load(file)
    
    cipher_suite = Fernet(encryption_key)
    
    decrypted_config = {
        key: cipher_suite.decrypt(value.encode()).decode()
        for key, value in encrypted_config.items()
    }
    return decrypted_config

def f_utils_create_a_connection_string_to_SQL_Server():
    
    encryption_key = load_encryption_key(PATH_CONFIG_KEY)
    config = load_config(PATH_CONFIG_JSON, encryption_key)
    
    # SQL-Server connection string with Server Authentication
    conn = pyodbc.connect(
        f'DRIVER={{SQL Server}};'
        f'SERVER={config["DB_HOST"]};'
        f'DATABASE={config["DB_NAME"]};'
        f'UID={config["DB_USER"]};'
        f'PWD={config["DB_PASSWORD"]};'
        'PORT=1433'
    )
    return conn

def f_utils_get_DB_HOST():
    # Get config and encryption-key
    encryption_key = load_encryption_key(PATH_CONFIG_KEY)
    config = load_config(PATH_CONFIG_JSON, encryption_key)
    # Get database_name
    database_host = config["DB_HOST"]
    return database_host

def f_utils_get_DB_NAME():
    # Get config and encryption-key
    encryption_key = load_encryption_key(PATH_CONFIG_KEY)
    config = load_config(PATH_CONFIG_JSON, encryption_key)
    # Get database_name
    database_name = config["DB_NAME"]
    return database_name

def f_utils_get_DB_USER_AND_DB_PASSWORD():
    # Get config and encryption-key
    encryption_key = load_encryption_key(PATH_CONFIG_KEY)
    config = load_config(PATH_CONFIG_JSON, encryption_key)
    # Get database_name
    DB_USER = config["DB_USER"]
    DB_PASSWORD = config["DB_PASSWORD"]
    return DB_USER, DB_PASSWORD

# Định dạng số theo yêu cầu
def f_utils_format_number(value):
    """ 
    Định dạng số theo chuẩn:
    - Số nguyên: 123,123
    - Số thực: 123,123.00
    - Các kiểu dữ liệu khác giữ nguyên
    """
    if isinstance(value, int):
        return f"{value:,}"
    elif isinstance(value, float):
        return f"{value:,.2f}"
    else:
        return value  # Trả về nguyên giá trị nếu không phải số

def f_utils_fetch_data_from_database(query):
    try:
        conn = f_utils_create_a_connection_string_to_SQL_Server()
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        # Đóng kết nối
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        return rows
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return []

def f_utils_fetch_data_from_database_with_quey_and_params(query, params_list=[]):
    try:
        conn = f_utils_create_a_connection_string_to_SQL_Server()
        cursor = conn.cursor()
        cursor.execute(query, params_list)
        rows = cursor.fetchall()
        # Đóng kết nối
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        return rows
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return []

def f_utils_sent_query_to_SQL(query):
    conn = f_utils_create_a_connection_string_to_SQL_Server()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute(query)
            # For non-SELECT queries, commit the transaction
            conn.commit()
            # print("Thành công", "Dữ liệu đã được cập nhật thành công!")
        except Exception as e:
            print(f"Error: {e}")
            print("Error at function: ", f_utils_get_current_function_name())
        finally:
            # Đóng kết nối
            if cursor:
                cursor.close()
            if conn:
                conn.close()


# ========================================================================================================================================================================
# inspect function name
# ========================================================================================================================================================================
# return current function name
def f_utils_get_current_function_name():
    func_name = inspect.currentframe().f_back.f_code.co_name
    # Lấy đường dẫn đầy đủ của file chứa hàm gọi
    file_name = inspect.currentframe().f_back.f_code.co_filename
    # return caller function name
    func_caller = inspect.currentframe().f_back.f_back.f_code.co_name
    return func_name, file_name, func_caller

# return caller function name
def f_utils_get_caller_function_name():
    return inspect.currentframe().f_back.f_back.f_code.co_name


def f_utils_check_duplicate(entry_to_check, database_name, table_name, column_name):
    value_to_check = entry_to_check.get().strip()
    
    if not value_to_check:
        messagebox.showwarning("Lỗi", "Vui lòng nhập giá trị cần kiểm tra!")
        return False

    try:
        conn = f_utils_create_a_connection_string_to_SQL_Server()
        cursor = conn.cursor()

        # Truy vấn kiểm tra sự tồn tại của số phiếu
        query = f"SELECT COUNT(*) FROM {database_name}.[dbo].{table_name} WHERE {column_name} = ? AND [XOA_SUA] = ''"
        cursor.execute(query, (value_to_check,))
        result = cursor.fetchone()[0]

        if result > 0:
            # Đóng kết nối
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            return False
        else:
            # Đóng kết nối
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            return True
        
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return False

def f_utils_check_exist(entry_to_check, database_name, table_name, column_name):
    value_to_check = entry_to_check.get().strip()
    
    if not value_to_check:
        messagebox.showwarning("Lỗi", "Vui lòng nhập giá trị cần kiểm tra!")
        return False

    try:
        conn = f_utils_create_a_connection_string_to_SQL_Server()
        cursor = conn.cursor()

        # Truy vấn kiểm tra sự tồn tại của số phiếu
        query = f"SELECT CASE WHEN EXISTS(SELECT 1 FROM {database_name}.[dbo].{table_name} WHERE {column_name} = ? AND [XOA_SUA] = '') THEN 1 ELSE 0 END"
        # print(query)
        # print(value_to_check)
        cursor.execute(query, (value_to_check,))
        result = cursor.fetchone()[0]

        # print("result: ", result)
        
        if result:
            # Đóng kết nối
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            return True
        else:
            messagebox.showinfo("Thông báo", f"Giá trị {value_to_check} chưa tồn tại. Vui lòng khởi tạo và thử lại!")
            # Đóng kết nối
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            return False
        
    except Exception as e:
        print(f"Error: {e}")
        print("Error at function: ", f_utils_get_current_function_name())
        return False
