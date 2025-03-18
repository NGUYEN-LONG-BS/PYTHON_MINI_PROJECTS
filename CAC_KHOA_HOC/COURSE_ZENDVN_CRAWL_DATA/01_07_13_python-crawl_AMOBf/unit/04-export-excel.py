import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font

mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        # password="password",
        database="python_crawl"
    )
cursor = mydb.cursor()

def execute_query(cursor, query, params=None):
    cursor.execute(query, params)
    result = cursor.fetchall()

    columns = [i[0] for i in cursor.description]

    return pd.DataFrame(result, columns=columns)

query_categories = """
    SELECT c.id, c.name, CONCAT('http://demo-php-bookstore.zendvn.com', c.link) AS link, COUNT(p.id) AS product_count
    FROM categories AS c
    LEFT JOIN products AS p ON c.id = p.category_id
    GROUP BY c.id
"""

query_products = """
    SELECT p.id, p.name, CONCAT('http://demo-php-bookstore.zendvn.com', p.link) AS link, CONCAT('http://demo-php-bookstore.zendvn.com', p.image_link) AS image_link, p.image_name, p.price_current, p.price_old, p.content, c.name AS category_name
    FROM products AS p
    LEFT JOIN categories AS c ON p.category_id = c.id
"""

df_categories = execute_query(cursor, query_categories)

df_categories.to_excel('categories.xlsx', index=False)

df_products = execute_query(cursor, query_products)

df_products.to_excel('products.xlsx', index=False)

cursor.close()
mydb.close()

wb = load_workbook('products.xlsx')
ws = wb.active

image_folder_path = './images/'

for index, row in df_products.iterrows():
    image = Image(f'{image_folder_path}{row['image_name']}')
    image.width = image.width * 0.75
    image.height = image.height * 0.75
    image.anchor = f'J{index+2}'

    ws.add_image(image)

ws.cell(row=1, column=10, value="image")

title_image = ws.cell(row=1, column=10)
title_image.border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
title_image.font = Font(bold=True)

wb.save('products_with_images.xlsx')